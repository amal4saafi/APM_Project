import openpyxl
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.core.mail import send_mail
from django.db.models import Count
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from openpyxl.styles import Font, PatternFill

Utilisateur = get_user_model()

from . import utils
from .decorators import admin_required
from .forms import (ApplicationForm, CertificatSSLForm, ContratForm,
                     DocumentForm, EnvironnementForm, ImportExcelForm,
                     NomDomaineForm, PasswordResetRequestForm, ProfilForm,
                     RessourcesGlobalesForm, RoleForm, SetNewPasswordForm,
                     UtilisateurCreateForm, UtilisateurEditForm)
from .models import (Application, CertificatSSL, Contrat, Document,
                      DocumentFichier, Environnement, HistoriqueAction, Log,
                      NomDomaine, Notification, PasswordResetToken,
                      ProfilUtilisateur, RessourcesGlobales, Role)


# Alias courts
def journaliser(request, action, modele='', objet=None, details=''):
    utils.journaliser(request, action, modele, objet, details)

def notifier(destinataires, titre, message, type_notif='info', action_url=''):
    utils.notifier(destinataires, titre, message, type_notif, action_url)

# ---------------------------------------------------------------------------
# Redirection post-login / dashboards
# ---------------------------------------------------------------------------

@login_required
def home_redirect(request):
    if request.user.is_admin_role:
        return redirect('dashboard_admin')
    if request.user.is_systeme_role:
        return redirect('dashboard_systeme')
    return redirect('dashboard_user')


@login_required
@admin_required
def dashboard_admin(request):
    """Dashboard de l'Admin : gestion des utilisateurs et des rôles uniquement."""
    import json
    from datetime import timedelta

    total_users = Utilisateur.objects.count()
    total_actifs = Utilisateur.objects.filter(is_active=True).count()
    total_inactifs = total_users - total_actifs
    total_roles = Role.objects.count()

    users_par_role = list(
        Utilisateur.objects.values('role__nom').annotate(total=Count('id')).order_by('-total')
    )
    roles_data = {
        'labels': [r['role__nom'] or 'Sans rôle' for r in users_par_role],
        'values': [r['total'] for r in users_par_role],
    }

    # Créations des 7 derniers jours
    today = timezone.now().date()
    jours = [today - timedelta(days=i) for i in range(6, -1, -1)]
    creations_values = []
    for j in jours:
        creations_values.append(Utilisateur.objects.filter(date_joined__date=j).count())
    creations_data = {
        'labels': [j.strftime('%d/%m') for j in jours],
        'values': creations_values,
    }

    derniers_users = Utilisateur.objects.select_related('role', 'profil').order_by('-date_joined')[:8]
    historique = HistoriqueAction.objects.select_related('utilisateur').order_by('-date_action')[:10]

    context = {
        'total_users': total_users,
        'total_actifs': total_actifs,
        'total_inactifs': total_inactifs,
        'total_roles': total_roles,
        'derniers_users': derniers_users,
        'historique': historique,
        'roles_data': json.dumps(roles_data),
        'creations_data': json.dumps(creations_data),
    }
    return render(request, 'apm/dashboard_admin.html', context)


@login_required
def dashboard_user(request):
    """Dashboard du Membre DSI : gestion complète du patrimoine applicatif."""
    import json
    from datetime import timedelta

    if not request.user.is_dsi_role:
        return redirect('home')

    utils.verifier_expirations_apps_externes()

    apps = Application.objects.filter(archive=False).order_by('-pk')[:12]
    envs = Environnement.objects.select_related('id_application').filter(archive=False).order_by('-pk')[:12]
    ssl_items = CertificatSSL.objects.select_related('id_environnement').filter(archive=False).order_by('-pk')[:6]
    all_envs = Environnement.objects.filter(archive=False)
    all_apps = Application.objects.filter(archive=False)

    critere_counts = list(all_apps.values('critere').annotate(total=Count('id_application')))
    critere_labels_map = {'eleve': 'Critique', 'normal': 'Normal', 'faible': 'Faible'}
    critique_data = {
        'labels': [critere_labels_map.get(c['critere'], c['critere']) for c in critere_counts],
        'values': [c['total'] for c in critere_counts],
    }

    ssl_actifs_qs = CertificatSSL.objects.filter(archive=False)
    ssl_valides = ssl_actifs_qs.exclude(statut__iexact='expiré').exclude(statut__icontains='30').count()
    ssl_bientot = ssl_actifs_qs.filter(statut__icontains='30').count()
    ssl_expires_count = ssl_actifs_qs.filter(statut__iexact='expiré').count()
    ssl_data = {'labels': ['Valides', 'Expire < 30j', 'Expirés'], 'values': [ssl_valides, ssl_bientot, ssl_expires_count]}

    env_types = list(all_envs.values('nom').annotate(total=Count('id_environnement')).order_by('-total')[:6])
    env_type_data = {'labels': [e['nom'] for e in env_types], 'values': [e['total'] for e in env_types]}

    today = timezone.now().date()
    jours = [today - timedelta(days=i) for i in range(6, -1, -1)]
    activite_values = [HistoriqueAction.objects.filter(date_action__date=j).count() for j in jours]
    activite_data = {'labels': [j.strftime('%d/%m') for j in jours], 'values': activite_values}

    historique = HistoriqueAction.objects.select_related('utilisateur').order_by('-date_action')[:10]

    context = {
        'apps': apps,
        'envs': envs,
        'ssl_items': ssl_items,
        'total_apps': all_apps.count(),
        'apps_critiques': all_apps.filter(critere=Application.CRITERE_ELEVE).count(),
        'total_env': all_envs.count(),
        'envs_sans_app': all_envs.filter(id_application__isnull=True).count(),
        'total_ssl': ssl_actifs_qs.count(),
        'ssl_expires': ssl_expires_count,
        'total_contrats': Contrat.objects.filter(archive=False).count(),
        'historique': historique,
        'critique_data': json.dumps(critique_data),
        'ssl_data': json.dumps(ssl_data),
        'env_type_data': json.dumps(env_type_data),
        'activite_data': json.dumps(activite_data),
    }
    return render(request, 'apm/dashboard_user.html', context)


@login_required
def dashboard_systeme(request):
    """Dashboard de l'Administrateur Système : consultation du patrimoine
    (applications / environnements / SSL disponibles) + gestion de la
    capacité totale CPU/RAM."""
    import json

    if not request.user.is_systeme_role:
        messages.error(request, "Accès réservé à l'Administrateur Système.")
        return redirect('home')

    ressources = RessourcesGlobales.get_solo()
    apps_qs = Application.objects.filter(archive=False)
    envs_qs = Environnement.objects.select_related('id_application').filter(archive=False)

    apps_actives = apps_qs.filter(statut=True).count()
    apps_inactives = apps_qs.count() - apps_actives
    app_statut_data = {'labels': ['Actives', 'Inactives'], 'values': [apps_actives, apps_inactives]}

    top_envs = envs_qs.order_by('-cpu')[:8]
    cpu_data = {'labels': [e.nom for e in top_envs], 'values': [e.cpu for e in top_envs]}
    ram_data = {'labels': [e.nom for e in top_envs], 'values': [e.ram for e in top_envs]}

    historique = HistoriqueAction.objects.select_related('utilisateur').order_by('-date_action')[:10]

    context = {
        'ressources': ressources,
        'total_apps': apps_qs.count(),
        'total_env': envs_qs.count(),
        'total_ssl': CertificatSSL.objects.filter(archive=False).count(),
        'ssl_expires': CertificatSSL.objects.filter(archive=False, statut__iexact='expiré').count(),
        'apps': apps_qs.order_by('nom')[:10],
        'envs': envs_qs.order_by('nom')[:10],
        'historique': historique,
        'app_statut_data': json.dumps(app_statut_data),
        'cpu_data': json.dumps(cpu_data),
        'ram_data': json.dumps(ram_data),
    }
    return render(request, 'apm/dashboard_systeme.html', context)


@login_required
def ressources_edit(request):
    """Édition de la capacité totale CPU/RAM — réservé à l'Administrateur Système."""
    if not request.user.is_systeme_role:
        messages.error(request, "Accès réservé à l'Administrateur Système.")
        return redirect('home')
    ressources = RessourcesGlobales.get_solo()
    if request.method == 'POST':
        form = RessourcesGlobalesForm(request.POST, instance=ressources)
        if form.is_valid():
            form.save()
            journaliser(request, "Modification capacité CPU/RAM", 'RessourcesGlobales', ressources)
            messages.success(request, "Capacité totale CPU/RAM mise à jour.")
            return redirect('dashboard_systeme')
    else:
        form = RessourcesGlobalesForm(instance=ressources)
    return render(request, 'apm/ressources_form.html', {'form': form, 'ressources': ressources})


# ---------------------------------------------------------------------------
# Gestion des utilisateurs (admin uniquement)
# ---------------------------------------------------------------------------

@login_required
@admin_required
def users_list(request):
    q = request.GET.get('q', '').strip()
    users = Utilisateur.objects.select_related('role', 'profil').all()
    if q:
        users = users.filter(username__icontains=q) | users.filter(email__icontains=q) \
            | users.filter(last_name__icontains=q)
    return render(request, 'apm/users_list.html', {'users': users.order_by('-date_joined'), 'q': q})


@login_required
@admin_required
def user_create(request):
    if request.method == 'POST':
        form = UtilisateurCreateForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data.get('email', '')
            if email and not utils.is_valid_email(email):
                form.add_error('email', 'Adresse email invalide.')
            else:
                user_obj = form.save()
                mdp_temp = form.cleaned_data.get('password1', '')
                utils.get_profil_or_create(user_obj)
                site_url = request.build_absolute_uri('/')
                ok, msg_email = utils.envoyer_email_creation_compte(user_obj, mdp_temp, site_url)
                journaliser(request, f'Création utilisateur : {user_obj.username}', 'Utilisateur', user_obj)
                notifier(
                    list(Utilisateur.objects.filter(role__type_role=Role.TYPE_ADMIN).exclude(pk=request.user.pk)),
                    'Nouvel utilisateur créé',
                    f"L'utilisateur « {user_obj.username} » a été créé par {request.user.username}.",
                    type_notif='info', action_url=reverse('users_list')
                )
                if email:
                    if ok:
                        messages.success(request, f'Utilisateur créé. Email de bienvenue envoyé à {email}.')
                    else:
                        messages.warning(request, f'Utilisateur créé. Email non envoyé : {msg_email}')
                else:
                    messages.success(request, 'Utilisateur créé avec succès.')
                return redirect('users_list')
    else:
        form = UtilisateurCreateForm()
    return render(request, 'apm/user_form.html', {'form': form, 'title': 'Nouvel utilisateur'})


@login_required
@admin_required
def user_edit(request, pk):
    user_obj = get_object_or_404(Utilisateur, pk=pk)
    if request.method == 'POST':
        form = UtilisateurEditForm(request.POST, instance=user_obj)
        if form.is_valid():
            form.save()
            messages.success(request, "Utilisateur mis à jour.")
            return redirect('users_list')
    else:
        form = UtilisateurEditForm(instance=user_obj)
    return render(request, 'apm/user_form.html', {'form': form, 'title': f"Modifier {user_obj.username}"})


@login_required
@admin_required
def user_toggle_active(request, pk):
    user_obj = get_object_or_404(Utilisateur, pk=pk)
    if user_obj == request.user:
        messages.error(request, "Vous ne pouvez pas désactiver votre propre compte.")
    else:
        user_obj.is_active = not user_obj.is_active
        user_obj.save(update_fields=['is_active'])
        etat = "activé" if user_obj.is_active else "désactivé"
        messages.success(request, f"Compte {user_obj.username} {etat}.")
    return redirect('users_list')


@login_required
@admin_required
def users_export_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Utilisateurs"

    headers = ['Username', 'Nom', 'Prénom', 'Email', 'Téléphone',
               'Département', 'Rôle', 'Actif', 'Date création']
    ws.append(headers)
    header_fill = PatternFill(start_color='4C6EF5', end_color='4C6EF5', fill_type='solid')
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill

    for u in Utilisateur.objects.select_related('role').all():
        ws.append([
            u.username, u.last_name, u.first_name, u.email, u.telephone,
            u.departement, u.role.nom if u.role else '', 'Oui' if u.is_active else 'Non',
            u.date_joined.strftime('%Y-%m-%d') if u.date_joined else '',
        ])

    for col in ws.columns:
        length = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(length + 2, 12), 40)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="utilisateurs_apm.xlsx"'
    wb.save(response)
    return response


@login_required
@admin_required
def users_import_excel(request):
    if request.method == 'POST':
        form = ImportExcelForm(request.POST, request.FILES)
        if form.is_valid():
            wb = openpyxl.load_workbook(request.FILES['fichier'], data_only=True)
            ws = wb.active
            created_list, updated_list, errors_list = [], [], []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                try:
                    username, nom, prenom, email = row[0], row[1] or '', row[2] or '', row[3] or ''
                    telephone = row[4] or '' if len(row) > 4 else ''
                    departement = row[5] or '' if len(row) > 5 else ''
                    role_nom = row[6] if len(row) > 6 else None
                    role_obj = Role.objects.filter(nom=role_nom).first() if role_nom else None

                    defaults = {
                        'last_name': nom, 'first_name': prenom, 'email': email,
                        'telephone': telephone, 'departement': departement,
                        'role': role_obj,
                    }
                    existing = Utilisateur.objects.filter(username=username).first()
                    a_change = False
                    if existing:
                        a_change = any(
                            getattr(existing, field) != value for field, value in defaults.items()
                        )

                    user_obj, was_created = Utilisateur.objects.update_or_create(
                        username=username, defaults=defaults
                    )
                    if was_created:
                        user_obj.set_unusable_password()
                        user_obj.save()
                        utils.get_profil_or_create(user_obj)
                        created_list.append(username)
                    elif a_change:
                        updated_list.append(username)
                    # Sinon : ligne identique à l'existant → ni ajout ni mise à jour comptabilisés.
                except Exception as e:
                    errors_list.append(str(row[0]))
            # If nothing changed during import, show concise message
            rapport = utils.build_import_report(created_list, updated_list, errors_list)
            if rapport['total'] == 0 and not rapport['errors']:
                messages.info(request, '0 ajout — aucune modification effectuée.')
            else:
                msg = rapport['resume']
                if rapport['created']:
                    msg += ' | Ajoutés : ' + ', '.join(rapport['created'][:5])
                    if len(rapport['created']) > 5:
                        msg += f" (+{len(rapport['created'])-5} autres)"
                if rapport['updated']:
                    msg += ' | Mis à jour : ' + ', '.join(rapport['updated'][:5])
                messages.success(request, f'Import terminé : {msg}')
            journaliser(request, f"Import Excel utilisateurs ({rapport['total']} lignes)", 'Utilisateur')
            return redirect('users_list')
    else:
        form = ImportExcelForm()
    return render(request, 'apm/user_import.html', {'form': form})


# ---------------------------------------------------------------------------
# Gestion des rôles (admin uniquement)
# ---------------------------------------------------------------------------

@login_required
@admin_required
def roles_list(request):
    return render(request, 'apm/roles_list.html', {'roles': Role.objects.all()})


@login_required
@admin_required
def role_create(request):
    if request.method == 'POST':
        form = RoleForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Rôle créé.")
            return redirect('roles_list')
    else:
        form = RoleForm()
    return render(request, 'apm/role_form.html', {'form': form, 'title': "Nouveau rôle"})


@login_required
@admin_required
def role_edit(request, pk):
    role_obj = get_object_or_404(Role, pk=pk)
    if request.method == 'POST':
        form = RoleForm(request.POST, instance=role_obj)
        if form.is_valid():
            form.save()
            messages.success(request, "Rôle mis à jour.")
            return redirect('roles_list')
    else:
        form = RoleForm(instance=role_obj)
    return render(request, 'apm/role_form.html', {'form': form, 'title': f"Modifier {role_obj.nom}"})


@login_required
@admin_required
def role_delete(request, pk):
    role_obj = get_object_or_404(Role, pk=pk)
    if request.method == 'POST':
        role_obj.delete()
        messages.success(request, "Rôle supprimé.")
        return redirect('roles_list')
    return render(request, 'apm/confirm_delete.html', {'obj': role_obj, 'back_url': 'roles_list'})


# ---------------------------------------------------------------------------
# CRUD Application (admin gère tout, user voit seulement les siennes en lecture)
# ---------------------------------------------------------------------------

@login_required
def applications_list(request):
    if not _patrimoine_view_allowed(request.user):
        messages.error(request, "Accès réservé à la DSI / à l'Administrateur Système.")
        return redirect('home')
    voir_archives = request.GET.get('archives') == '1'
    if not voir_archives:
        utils.verifier_expirations_apps_externes()
    apps = Application.objects.select_related('id_user').filter(archive=voir_archives)
    return render(request, 'apm/applications_list.html', {
        'apps': apps.order_by('nom'), 'voir_archives': voir_archives,
    })


@login_required
def application_create(request):
    if not _dsi_allowed(request.user):
        messages.error(request, "Accès réservé à la DSI.")
        return redirect('applications_list')
    env_id = request.GET.get('env')
    linked_env = Environnement.objects.filter(pk=env_id, id_application__isnull=True).first() if env_id else None
    if request.method == 'POST':
        form = ApplicationForm(request.POST)
        if form.is_valid():
            app_obj = form.save()
            if linked_env:
                linked_env.id_application = app_obj
                linked_env.save(update_fields=['id_application'])
                messages.success(request, f"Application créée et liée à l'environnement « {linked_env.nom} ».")
            else:
                messages.success(request, "Application créée.")
            if form.cleaned_data.get('exige_ssl'):
                return redirect(reverse('ssl_create') + f"?app={app_obj.pk}")
            return redirect('applications_list')
    else:
        form = ApplicationForm()
    return render(request, 'apm/application_form.html', {
        'form': form, 'title': "Nouvelle application", 'linked_env': linked_env,
    })


@login_required
def application_edit(request, pk):
    app_obj = get_object_or_404(Application, pk=pk)
    if not _dsi_allowed(request.user):
        messages.error(request, "Accès réservé à la DSI.")
        return redirect('applications_list')
    if request.method == 'POST':
        form = ApplicationForm(request.POST, instance=app_obj)
        if form.is_valid():
            app_obj = form.save()
            messages.success(request, "Application mise à jour.")
            if form.cleaned_data.get('exige_ssl'):
                return redirect(reverse('ssl_create') + f"?app={app_obj.pk}")
            return redirect('applications_list')
    else:
        form = ApplicationForm(instance=app_obj)
    return render(request, 'apm/application_form.html', {'form': form, 'title': f"Modifier {app_obj.nom}"})


@login_required
def application_delete(request, pk):
    """Archive l'application (aucune suppression définitive des données)."""
    app_obj = get_object_or_404(Application, pk=pk)
    if not _dsi_allowed(request.user):
        messages.error(request, "Accès réservé aux administrateurs/DSI.")
        return redirect('applications_list')
    if request.method == 'POST':
        app_obj.archive = True
        app_obj.save(update_fields=['archive'])
        journaliser(request, f"Archivage application : {app_obj.nom}", 'Application', app_obj)
        messages.success(request, "Application archivée.")
        return redirect('applications_list')
    return render(request, 'apm/confirm_delete.html', {'obj': app_obj, 'back_url': 'applications_list', 'action_label': 'archiver'})


@login_required
def application_restore(request, pk):
    app_obj = get_object_or_404(Application, pk=pk)
    if not _dsi_allowed(request.user):
        messages.error(request, "Accès réservé aux administrateurs/DSI.")
        return redirect('applications_list')
    app_obj.archive = False
    app_obj.save(update_fields=['archive'])
    journaliser(request, f"Restauration application : {app_obj.nom}", 'Application', app_obj)
    messages.success(request, "Application restaurée.")
    return redirect(request.META.get('HTTP_REFERER') or reverse('applications_list') + '?archives=1')


# ---------------------------------------------------------------------------
# CRUD Environnement
# ---------------------------------------------------------------------------

@login_required
def environnements_list(request):
    if not _patrimoine_view_allowed(request.user):
        messages.error(request, "Accès réservé à la DSI / à l'Administrateur Système.")
        return redirect('home')
    voir_archives = request.GET.get('archives') == '1'
    envs = Environnement.objects.select_related('id_application', 'id_serveur').filter(archive=voir_archives)
    return render(request, 'apm/environnements_list.html', {
        'envs': envs.order_by('id_application__nom', 'nom'), 'voir_archives': voir_archives,
    })


@login_required
def environnement_create(request):
    if not _dsi_allowed(request.user):
        messages.error(request, "Accès réservé à la DSI.")
        return redirect('environnements_list')
    if request.method == 'POST':
        form = EnvironnementForm(request.POST)
        if form.is_valid():
            env_obj = form.save()
            # Étape 1 : si l'environnement n'est lié à aucune application, on
            # redirige vers la création d'une application liée à cet environnement.
            if not env_obj.id_application_id:
                messages.info(request, "Environnement créé. Il faut maintenant lui associer une application.")
                return redirect(reverse('application_create') + f"?env={env_obj.pk}")
            messages.success(request, "Environnement créé.")
            # Étape 2 : si l'environnement exige un certificat SSL, on le propose ensuite.
            if env_obj.exige_ssl:
                return redirect(reverse('ssl_create') + f"?env={env_obj.pk}")
            return redirect('environnements_list')
    else:
        form = EnvironnementForm()
    return render(request, 'apm/environnement_form.html', {'form': form, 'title': "Nouvel environnement"})


@login_required
def environnement_edit(request, pk):
    env_obj = get_object_or_404(Environnement, pk=pk)
    if not _dsi_allowed(request.user):
        messages.error(request, "Accès réservé à la DSI.")
        return redirect('environnements_list')
    if request.method == 'POST':
        form = EnvironnementForm(request.POST, instance=env_obj)
        if form.is_valid():
            env_obj = form.save()
            if not env_obj.id_application_id:
                messages.info(request, "Il faut associer une application à cet environnement.")
                return redirect(reverse('application_create') + f"?env={env_obj.pk}")
            messages.success(request, "Environnement mis à jour.")
            if env_obj.exige_ssl:
                return redirect(reverse('ssl_create') + f"?env={env_obj.pk}")
            return redirect('environnements_list')
    else:
        form = EnvironnementForm(instance=env_obj)
    return render(request, 'apm/environnement_form.html', {'form': form, 'title': f"Modifier {env_obj.nom}"})


@login_required
def environnement_delete(request, pk):
    env_obj = get_object_or_404(Environnement, pk=pk)
    if not _dsi_allowed(request.user):
        messages.error(request, "Accès réservé à la DSI.")
        return redirect('environnements_list')
    if request.method == 'POST':
        env_obj.archive = True
        env_obj.save(update_fields=['archive'])
        journaliser(request, f"Archivage environnement : {env_obj.nom}", 'Environnement', env_obj)
        messages.success(request, "Environnement archivé.")
        return redirect('environnements_list')
    return render(request, 'apm/confirm_delete.html', {'obj': env_obj, 'back_url': 'environnements_list', 'action_label': 'archiver'})


@login_required
def environnement_restore(request, pk):
    env_obj = get_object_or_404(Environnement, pk=pk)
    if not _dsi_allowed(request.user):
        messages.error(request, "Accès réservé à la DSI.")
        return redirect('environnements_list')
    env_obj.archive = False
    env_obj.save(update_fields=['archive'])
    journaliser(request, f"Restauration environnement : {env_obj.nom}", 'Environnement', env_obj)
    messages.success(request, "Environnement restauré.")
    return redirect(request.META.get('HTTP_REFERER') or reverse('environnements_list') + '?archives=1')


# ---------------------------------------------------------------------------
# Helpers de permissions — 3 rôles : admin / admin_systeme / membre_dsi
# ---------------------------------------------------------------------------


def _dsi_allowed(user):
    """Gestion complète du patrimoine (CRUD) : réservé au Membre DSI."""
    return bool(user.is_authenticated and user.is_dsi_role)


def _ssl_renew_allowed(user):
    """Renouvellement effectif d'un certificat SSL expiré : réservé à
    l'Administrateur Système."""
    return bool(user.is_authenticated and user.is_systeme_role)


def _systeme_allowed(user):
    """Gestion de la capacité CPU/RAM : réservé à l'Administrateur Système."""
    return bool(user.is_authenticated and user.is_systeme_role)


def _patrimoine_view_allowed(user):
    """Consultation des listes (apps/env/ssl/domaines) : DSI (CRUD) ou
    Administrateur Système (lecture seule)."""
    return bool(user.is_authenticated and (user.is_dsi_role or user.is_systeme_role))


@login_required
def ssl_list(request):
    if not _patrimoine_view_allowed(request.user):
        messages.error(request, "Accès réservé aux équipes techniques.")
        return redirect('home')
    voir_archives = request.GET.get('archives') == '1'
    items = CertificatSSL.objects.select_related('id_environnement').filter(archive=voir_archives)
    expires_count = sum(1 for it in items if it.est_expire)
    return render(request, 'apm/ssl_list.html', {
        'items': items, 'voir_archives': voir_archives, 'expires_count': expires_count,
    })


@login_required
def ssl_create(request):
    if not _dsi_allowed(request.user):
        messages.error(request, "Accès réservé aux équipes techniques.")
        return redirect('ssl_list')
    if request.method == 'POST':
        form = CertificatSSLForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Certificat SSL créé.")
            return redirect('ssl_list')
    else:
        env_id = request.GET.get('env')
        app_id = request.GET.get('app')
        if app_id:
            # If an app is provided, try to prefill or limit environments
            app_obj = Application.objects.filter(pk=app_id).first()
            if app_obj:
                envs = app_obj.environnements.all()
                if envs.count() == 1:
                    form = CertificatSSLForm(app_id=app_id, initial={'id_environnement': envs.first().pk})
                else:
                    form = CertificatSSLForm(app_id=app_id)
            else:
                form = CertificatSSLForm()
        elif env_id:
            form = CertificatSSLForm(initial={'id_environnement': env_id})
        else:
            form = CertificatSSLForm()
    return render(request, 'apm/ssl_form.html', {'form': form, 'title': 'Nouveau certificat SSL'})


@login_required
def ssl_edit(request, pk):
    cert = get_object_or_404(CertificatSSL, pk=pk)
    est_renouvellement = _ssl_renew_allowed(request.user) and cert.est_expire and not _dsi_allowed(request.user)
    if not (_dsi_allowed(request.user) or est_renouvellement):
        if _ssl_renew_allowed(request.user):
            messages.error(request, "L'Administrateur Système ne peut renouveler que les certificats SSL expirés.")
        else:
            messages.error(request, "Accès réservé aux équipes techniques.")
        return redirect('ssl_list')
    if request.method == 'POST':
        form = CertificatSSLForm(request.POST, instance=cert)
        if form.is_valid():
            form.save()
            if est_renouvellement:
                messages.success(request, "Certificat SSL renouvelé.")
                journaliser(request, f"Renouvellement certificat SSL : {cert.domaine}", 'CertificatSSL', cert)
            else:
                messages.success(request, "Certificat SSL mis à jour.")
            return redirect('ssl_list')
    else:
        form = CertificatSSLForm(instance=cert)
    titre = f"Renouveler {cert.domaine}" if est_renouvellement else f"Modifier {cert.domaine}"
    return render(request, 'apm/ssl_form.html', {'form': form, 'title': titre})


@login_required
def ssl_demande_renouvellement(request, pk):
    """Le Membre DSI ne renouvelle pas lui-même un certificat SSL expiré :
    il envoie une demande de renouvellement à l'Administrateur Système."""
    cert = get_object_or_404(CertificatSSL, pk=pk)
    if not _dsi_allowed(request.user):
        messages.error(request, "Accès réservé aux équipes techniques.")
        return redirect('ssl_list')
    if request.method == 'POST':
        if not cert.est_expire:
            messages.error(request, "Ce certificat n'est pas expiré.")
            return redirect('ssl_list')
        utils.notifier_tous_admins(
            titre="Demande de renouvellement SSL",
            message=(
                f"{request.user.get_full_name() or request.user.username} demande le "
                f"renouvellement du certificat SSL « {cert.domaine} » "
                f"(expiré depuis {abs(cert.jours_restants)} jour(s))."
            ),
            type_notif='warning',
            action_url=reverse('ssl_edit', args=[cert.pk]),
        )
        journaliser(request, f"Demande de renouvellement SSL : {cert.domaine}", 'CertificatSSL', cert)
        messages.success(request, "Demande de renouvellement envoyée à l'Administrateur Système.")
    return redirect('ssl_list')


@login_required
def ssl_delete(request, pk):
    cert = get_object_or_404(CertificatSSL, pk=pk)
    if not _dsi_allowed(request.user):
        messages.error(request, "Accès réservé aux équipes techniques.")
        return redirect('ssl_list')
    if request.method == 'POST':
        cert.archive = True
        cert.save(update_fields=['archive'])
        journaliser(request, f"Archivage certificat SSL : {cert.domaine}", 'CertificatSSL', cert)
        messages.success(request, "Certificat SSL archivé.")
        return redirect('ssl_list')
    return render(request, 'apm/confirm_delete.html', {'obj': cert, 'back_url': 'ssl_list', 'action_label': 'archiver'})


@login_required
def ssl_restore(request, pk):
    cert = get_object_or_404(CertificatSSL, pk=pk)
    if not _dsi_allowed(request.user):
        messages.error(request, "Accès réservé aux équipes techniques.")
        return redirect('ssl_list')
    cert.archive = False
    cert.save(update_fields=['archive'])
    journaliser(request, f"Restauration certificat SSL : {cert.domaine}", 'CertificatSSL', cert)
    messages.success(request, "Certificat SSL restauré.")
    return redirect(request.META.get('HTTP_REFERER') or reverse('ssl_list') + '?archives=1')


@login_required
def noms_domaine_list(request):
    if not _patrimoine_view_allowed(request.user):
        messages.error(request, "Accès réservé aux équipes techniques.")
        return redirect('home')
    voir_archives = request.GET.get('archives') == '1'
    items = NomDomaine.objects.select_related('id_application').filter(archive=voir_archives)
    return render(request, 'apm/noms_domaine_list.html', {'items': items, 'voir_archives': voir_archives})


@login_required
def noms_domaine_create(request):
    if not _dsi_allowed(request.user):
        messages.error(request, "Accès réservé aux équipes techniques.")
        return redirect('noms_domaine_list')
    if request.method == 'POST':
        form = NomDomaineForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Nom de domaine créé.")
            return redirect('noms_domaine_list')
    else:
        form = NomDomaineForm()
    return render(request, 'apm/nom_domaine_form.html', {'form': form, 'title': 'Nouveau nom de domaine'})


@login_required
def noms_domaine_edit(request, pk):
    item = get_object_or_404(NomDomaine, pk=pk)
    if not _dsi_allowed(request.user):
        messages.error(request, "Accès réservé aux équipes techniques.")
        return redirect('noms_domaine_list')
    if request.method == 'POST':
        form = NomDomaineForm(request.POST, instance=item)
        if form.is_valid():
            form.save()
            messages.success(request, "Nom de domaine mis à jour.")
            return redirect('noms_domaine_list')
    else:
        form = NomDomaineForm(instance=item)
    return render(request, 'apm/nom_domaine_form.html', {'form': form, 'title': f"Modifier {item.nom_domaine}"})


@login_required
def noms_domaine_delete(request, pk):
    item = get_object_or_404(NomDomaine, pk=pk)
    if not _dsi_allowed(request.user):
        messages.error(request, "Accès réservé aux équipes techniques.")
        return redirect('noms_domaine_list')
    if request.method == 'POST':
        item.archive = True
        item.save(update_fields=['archive'])
        journaliser(request, f"Archivage nom de domaine : {item.nom_domaine}", 'NomDomaine', item)
        messages.success(request, "Nom de domaine archivé.")
        return redirect('noms_domaine_list')
    return render(request, 'apm/confirm_delete.html', {'obj': item, 'back_url': 'noms_domaine_list', 'action_label': 'archiver'})


@login_required
def noms_domaine_restore(request, pk):
    item = get_object_or_404(NomDomaine, pk=pk)
    if not _dsi_allowed(request.user):
        messages.error(request, "Accès réservé aux équipes techniques.")
        return redirect('noms_domaine_list')
    item.archive = False
    item.save(update_fields=['archive'])
    journaliser(request, f"Restauration nom de domaine : {item.nom_domaine}", 'NomDomaine', item)
    messages.success(request, "Nom de domaine restauré.")
    return redirect(request.META.get('HTTP_REFERER') or reverse('noms_domaine_list') + '?archives=1')


@login_required
def contrats_list(request):
    if not _patrimoine_view_allowed(request.user):
        messages.error(request, "Accès réservé aux équipes techniques.")
        return redirect('home')
    voir_archives = request.GET.get('archives') == '1'
    items = Contrat.objects.select_related('id_application').filter(archive=voir_archives)
    return render(request, 'apm/contrats_list.html', {'items': items, 'voir_archives': voir_archives})


@login_required
def contrats_create(request):
    if not _dsi_allowed(request.user):
        messages.error(request, "Accès réservé aux équipes techniques.")
        return redirect('contrats_list')
    if request.method == 'POST':
        form = ContratForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Contrat créé.")
            return redirect('contrats_list')
    else:
        form = ContratForm()
    return render(request, 'apm/contrat_form.html', {'form': form, 'title': 'Nouveau contrat'})


@login_required
def contrats_edit(request, pk):
    item = get_object_or_404(Contrat, pk=pk)
    if not _dsi_allowed(request.user):
        messages.error(request, "Accès réservé aux équipes techniques.")
        return redirect('contrats_list')
    if request.method == 'POST':
        form = ContratForm(request.POST, instance=item)
        if form.is_valid():
            form.save()
            messages.success(request, "Contrat mis à jour.")
            return redirect('contrats_list')
    else:
        form = ContratForm(instance=item)
    return render(request, 'apm/contrat_form.html', {'form': form, 'title': f"Modifier {item.numero_contrat}"})


@login_required
def contrats_delete(request, pk):
    item = get_object_or_404(Contrat, pk=pk)
    if not _dsi_allowed(request.user):
        messages.error(request, "Accès réservé aux équipes techniques.")
        return redirect('contrats_list')
    if request.method == 'POST':
        item.archive = True
        item.save(update_fields=['archive'])
        journaliser(request, f"Archivage contrat : {item.numero_contrat}", 'Contrat', item)
        messages.success(request, "Contrat archivé.")
        return redirect('contrats_list')
    return render(request, 'apm/confirm_delete.html', {'obj': item, 'back_url': 'contrats_list', 'action_label': 'archiver'})


@login_required
def contrats_restore(request, pk):
    item = get_object_or_404(Contrat, pk=pk)
    if not _dsi_allowed(request.user):
        messages.error(request, "Accès réservé aux équipes techniques.")
        return redirect('contrats_list')
    item.archive = False
    item.save(update_fields=['archive'])
    journaliser(request, f"Restauration contrat : {item.numero_contrat}", 'Contrat', item)
    messages.success(request, "Contrat restauré.")
    return redirect(request.META.get('HTTP_REFERER') or reverse('contrats_list') + '?archives=1')


@login_required
def documents_list(request):
    if not _patrimoine_view_allowed(request.user):
        messages.error(request, "Accès réservé aux équipes techniques.")
        return redirect('home')
    voir_archives = request.GET.get('archives') == '1'
    items = Document.objects.select_related('id_application').filter(archive=voir_archives)
    return render(request, 'apm/documents_list.html', {'items': items, 'voir_archives': voir_archives})


EXTS_OK_DOC = {'pdf', 'docx', 'xlsx', 'png', 'jpg', 'jpeg'}
TYPE_MAP_DOC = {'pdf': 'pdf', 'docx': 'docx', 'xlsx': 'xlsx', 'png': 'png', 'jpg': 'jpg', 'jpeg': 'jpg'}


@login_required
def documents_create(request):
    if not _dsi_allowed(request.user):
        messages.error(request, "Accès réservé aux équipes techniques.")
        return redirect('documents_list')
    if request.method == 'POST':
        form = DocumentForm(request.POST, request.FILES)
        if form.is_valid():
            doc = form.save()
            journaliser(request, f"Création document : {doc.nom_fichier}", 'Document', doc)

            fichiers = request.FILES.getlist('fichiers')
            nb_ok, nb_rejetes = 0, 0
            for fichier in fichiers:
                ext = fichier.name.rsplit('.', 1)[-1].lower() if '.' in fichier.name else 'autre'
                if ext not in EXTS_OK_DOC:
                    nb_rejetes += 1
                    continue
                taille_ko = max(1, fichier.size // 1024)
                DocumentFichier.objects.create(
                    document=doc, fichier=fichier, nom_original=fichier.name,
                    type_fichier=TYPE_MAP_DOC.get(ext, 'autre'), taille_ko=taille_ko,
                    uploade_par=request.user,
                )
                nb_ok += 1

            if nb_ok:
                journaliser(request, f"Import de {nb_ok} fichier(s) au document : {doc.nom_fichier}", 'DocumentFichier', doc)
                msg = f"Document créé avec {nb_ok} fichier(s) importé(s)."
                if nb_rejetes:
                    msg += f" {nb_rejetes} fichier(s) ignoré(s) (format non autorisé)."
                messages.success(request, msg)
            else:
                if nb_rejetes:
                    messages.warning(request, f"Document créé, mais {nb_rejetes} fichier(s) ignoré(s) (format non autorisé).")
                else:
                    messages.success(request, "Document créé. Vous pouvez maintenant y attacher des fichiers.")
            return redirect('documents_edit', pk=doc.pk)
    else:
        form = DocumentForm()
    return render(request, 'apm/document_form.html', {'form': form, 'title': 'Nouveau document'})


@login_required
def documents_edit(request, pk):
    item = get_object_or_404(Document, pk=pk)
    if not _dsi_allowed(request.user):
        messages.error(request, "Accès réservé aux équipes techniques.")
        return redirect('documents_list')
    if request.method == 'POST':
        form = DocumentForm(request.POST, instance=item)
        if form.is_valid():
            form.save()
            journaliser(request, f"Modification document : {item.nom_fichier}", 'Document', item)
            messages.success(request, "Document mis à jour.")
            return redirect('documents_list')
    else:
        form = DocumentForm(instance=item)
    return render(request, 'apm/document_form.html', {
        'form': form, 'title': f"Modifier {item.nom_fichier}", 'doc': item,
        'fichiers': item.fichiers.all().order_by('-uploade_le'),
    })


@login_required
def documents_delete(request, pk):
    item = get_object_or_404(Document, pk=pk)
    if not _dsi_allowed(request.user):
        messages.error(request, "Accès réservé aux équipes techniques.")
        return redirect('documents_list')
    if request.method == 'POST':
        item.archive = True
        item.save(update_fields=['archive'])
        journaliser(request, f"Archivage document : {item.nom_fichier}", 'Document', item)
        messages.success(request, "Document archivé.")
        return redirect('documents_list')
    return render(request, 'apm/confirm_delete.html', {'obj': item, 'back_url': 'documents_list', 'action_label': 'archiver'})


@login_required
def documents_restore(request, pk):
    item = get_object_or_404(Document, pk=pk)
    if not _dsi_allowed(request.user):
        messages.error(request, "Accès réservé aux équipes techniques.")
        return redirect('documents_list')
    item.archive = False
    item.save(update_fields=['archive'])
    journaliser(request, f"Restauration document : {item.nom_fichier}", 'Document', item)
    messages.success(request, "Document restauré.")
    return redirect(request.META.get('HTTP_REFERER') or reverse('documents_list') + '?archives=1')
    return render(request, 'apm/confirm_delete.html', {'obj': item, 'back_url': 'documents_list'})


@login_required
def documentation_page(request):
    if not _patrimoine_view_allowed(request.user):
        messages.error(request, "Accès réservé aux équipes techniques.")
        return redirect('home')
    docs = Document.objects.select_related('id_application').prefetch_related('fichiers').order_by('-date_upload')
    categories = list(docs.exclude(categorie='').values_list('categorie', flat=True).distinct())
    return render(request, 'apm/documentation.html', {
        'docs': docs,
        'categories': categories,
        'applications': Application.objects.order_by('nom'),
    })


@login_required
def documentation_quick_upload(request):
    """Upload rapide depuis l'espace Documentation : crée le Document + fichier en une étape."""
    if not _dsi_allowed(request.user):
        messages.error(request, "Accès réservé à la DSI.")
        return redirect('documentation_page')
    if request.method == 'POST':
        fichier = request.FILES.get('fichier')
        categorie = request.POST.get('categorie', '').strip()
        app_id = request.POST.get('id_application') or None
        description = request.POST.get('description', '').strip()

        if not fichier:
            messages.error(request, "Aucun fichier sélectionné.")
            return redirect('documentation_page')

        ext = fichier.name.rsplit('.', 1)[-1].lower() if '.' in fichier.name else 'autre'
        EXTS_OK = {'pdf', 'docx', 'xlsx', 'png', 'jpg', 'jpeg'}
        if ext not in EXTS_OK:
            messages.error(request, f"Extension « .{ext} » non autorisée. Formats : PDF, DOCX, XLSX, PNG, JPG.")
            return redirect('documentation_page')

        app_obj = Application.objects.filter(pk=app_id).first() if app_id else Application.objects.first()
        if not app_obj:
            messages.error(request, "Créez d'abord une application avant d'ajouter de la documentation.")
            return redirect('documentation_page')

        type_map = {'pdf': 'pdf', 'docx': 'docx', 'xlsx': 'xlsx', 'png': 'png', 'jpg': 'jpg', 'jpeg': 'jpg'}
        doc = Document.objects.create(
            nom_fichier=fichier.name,
            type_fichier=ext,
            categorie=categorie,
            description=description,
            id_application=app_obj,
        )
        taille_ko = max(1, fichier.size // 1024)
        DocumentFichier.objects.create(
            document=doc, fichier=fichier, nom_original=fichier.name,
            type_fichier=type_map.get(ext, 'autre'), taille_ko=taille_ko, uploade_par=request.user,
        )
        journaliser(request, f"Upload documentation : {fichier.name}", 'Document', doc)
        notifier(
            list(Utilisateur.objects.filter(role__type_role=Role.TYPE_MEMBRE_DSI).exclude(pk=request.user.pk)),
            "Nouveau document ajouté",
            f"« {fichier.name} » a été ajouté à la documentation ({app_obj.nom}).",
            type_notif='success', action_url=reverse('documentation_page')
        )
        messages.success(request, f"Fichier « {fichier.name} » ajouté à la documentation ({taille_ko} Ko).")
    return redirect('documentation_page')


@login_required
def rapports_page(request):
    if not _patrimoine_view_allowed(request.user):
        messages.error(request, "Accès réservé aux équipes techniques.")
        return redirect('home')
    return render(request, 'apm/rapports.html')


# ===========================================================================
# Mot de passe oublié — flux complet
# ===========================================================================

def _send_reset_email(request, user, token_obj):
    """Envoie l'email de réinitialisation (console en dev, SMTP en prod)."""
    reset_url = request.build_absolute_uri(
        reverse('password_reset_confirm', kwargs={'token': token_obj.token})
    )
    sujet = "Réinitialisation de votre mot de passe — APM TOPNET"
    corps = (
        f"Bonjour {user.get_full_name() or user.username},\n\n"
        f"Vous avez demandé la réinitialisation de votre mot de passe sur la plateforme APM TOPNET.\n\n"
        f"Cliquez sur le lien ci-dessous pour choisir un nouveau mot de passe :\n\n"
        f"{reset_url}\n\n"
        f"Ce lien est valable 24 heures. Si vous n'avez pas fait cette demande, ignorez cet email.\n\n"
        f"L'équipe APM TOPNET"
    )
    try:
        send_mail(sujet, corps, settings.DEFAULT_FROM_EMAIL, [user.email], fail_silently=False)
        return True
    except Exception as exc:
        return False


def password_reset_request(request):
    """Étape 1 : Demande de réinitialisation (email ou nom d'utilisateur)."""
    if request.user.is_authenticated:
        return redirect('home')

    if request.method == 'POST':
        form = PasswordResetRequestForm(request.POST)
        if form.is_valid():
            user = form.cleaned_data['user']
            token_obj = PasswordResetToken.create_for_user(user)

            if user.email:
                sent = _send_reset_email(request, user, token_obj)
                if sent:
                    messages.success(
                        request,
                        f"Un email a été envoyé à {user.email}. "
                        "Vérifiez votre boîte (et les spams)."
                    )
                else:
                    # Email échoué → on affiche le lien en fallback (dev/debug)
                    reset_url = request.build_absolute_uri(
                        reverse('password_reset_confirm', kwargs={'token': token_obj.token})
                    )
                    messages.warning(
                        request,
                        f"L'envoi d'email a échoué. Lien de reset (à usage unique) : {reset_url}"
                    )
            else:
                # Pas d'email configuré → on affiche le lien directement
                reset_url = request.build_absolute_uri(
                    reverse('password_reset_confirm', kwargs={'token': token_obj.token})
                )
                messages.info(
                    request,
                    f"Aucun email n'est configuré pour ce compte. "
                    f"Transmettez ce lien à l'utilisateur (valable 24h) : {reset_url}"
                )
            return redirect('password_reset_done')
    else:
        form = PasswordResetRequestForm()

    return render(request, 'registration/password_reset_request.html', {'form': form})


def password_reset_done(request):
    """Page de confirmation après soumission de la demande."""
    if request.user.is_authenticated:
        return redirect('home')
    return render(request, 'registration/password_reset_done.html')


def password_reset_confirm(request, token):
    """Étape 2 : Saisie du nouveau mot de passe via le lien tokenisé."""
    token_obj = PasswordResetToken.objects.filter(token=token).select_related('user').first()

    if not token_obj or not token_obj.is_valid:
        return render(request, 'registration/password_reset_invalid.html', {
            'raison': "Ce lien de réinitialisation est invalide ou a expiré."
        })

    if request.method == 'POST':
        form = SetNewPasswordForm(request.POST)
        if form.is_valid():
            user = token_obj.user
            user.set_password(form.cleaned_data['nouveau_mdp'])
            user.save(update_fields=['password'])
            token_obj.utilise = True
            token_obj.save(update_fields=['utilise'])
            messages.success(request, "Mot de passe mis à jour avec succès. Vous pouvez vous connecter.")
            return redirect('password_reset_complete')
    else:
        form = SetNewPasswordForm()

    return render(request, 'registration/password_reset_confirm.html', {
        'form': form,
        'token': token,
        'user': token_obj.user,
    })


def password_reset_complete(request):
    """Page finale après changement réussi."""
    return render(request, 'registration/password_reset_complete.html')


# ===========================================================================
# PROFIL UTILISATEUR — commun aux 3 rôles
# ===========================================================================

@login_required
def profil_view(request):
    profil = utils.get_profil_or_create(request.user)
    mes_notifs = Notification.objects.filter(destinataire=request.user).order_by('-cree_le')[:20]
    mes_actions = HistoriqueAction.objects.filter(utilisateur=request.user).order_by('-date_action')[:30]
    return render(request, 'apm/profil.html', {
        'profil': profil,
        'mes_notifs': mes_notifs,
        'mes_actions': mes_actions,
    })


@login_required
def profil_edit(request):
    profil = utils.get_profil_or_create(request.user)
    from .forms import ProfilForm
    if request.method == 'POST':
        form = ProfilForm(request.POST, request.FILES, instance=profil, user=request.user)
        if form.is_valid():
            form.save()
            journaliser(request, 'Modification du profil', 'ProfilUtilisateur', profil)
            messages.success(request, 'Profil mis à jour.')
            return redirect('profil')
    else:
        form = ProfilForm(instance=profil, user=request.user)
    return render(request, 'apm/profil_form.html', {'form': form, 'profil': profil})


@login_required
def profil_reset_mdp(request):
    """Chaque utilisateur peut réinitialiser son propre mot de passe depuis son profil."""
    from .forms import SetNewPasswordForm
    if request.method == 'POST':
        form = SetNewPasswordForm(request.POST)
        if form.is_valid():
            request.user.set_password(form.cleaned_data['nouveau_mdp'])
            request.user.save(update_fields=['password'])
            journaliser(request, 'Changement de mot de passe', 'Utilisateur', request.user)
            messages.success(request, 'Mot de passe changé avec succès. Reconnectez-vous.')
            return redirect('login')
    else:
        form = SetNewPasswordForm()
    return render(request, 'apm/profil_reset_mdp.html', {'form': form})


# ===========================================================================
# NOTIFICATIONS
# ===========================================================================

@login_required
def notification_mark_read(request, pk):
    n = get_object_or_404(Notification, pk=pk, destinataire=request.user)
    n.lu = True
    n.save(update_fields=['lu'])
    if n.action_url:
        return redirect(n.action_url)
    return redirect(request.META.get('HTTP_REFERER', 'home'))


@login_required
def notifications_mark_all_read(request):
    Notification.objects.filter(destinataire=request.user, lu=False).update(lu=True)
    messages.success(request, 'Toutes les notifications marquées comme lues.')
    return redirect(request.META.get('HTTP_REFERER', 'home'))


@login_required
def api_notifications(request):
    """API JSON — nombre de notifications non lues + liste."""
    import json
    from django.http import JsonResponse
    notifs = Notification.objects.filter(destinataire=request.user, lu=False).order_by('-cree_le')[:10]
    data = {
        'count': notifs.count(),
        'items': [
            {
                'id': n.pk,
                'type': n.type_notif,
                'titre': n.titre,
                'message': n.message,
                'date': n.cree_le.strftime('%d/%m/%Y %H:%M'),
                'url': n.action_url or '',
            }
            for n in notifs
        ]
    }
    return JsonResponse(data)


# ===========================================================================
# HISTORIQUE GLOBAL
# ===========================================================================

@login_required
def historique_page(request):
    """Page historique : filtrable par date et utilisateur."""
    import json
    from datetime import timedelta

    qs = HistoriqueAction.objects.select_related('utilisateur').order_by('-date_action')
    # Filtres
    date_filtre = request.GET.get('date', '')
    user_filtre = request.GET.get('user', '')
    modele_filtre = request.GET.get('modele', '')
    if date_filtre:
        qs = qs.filter(date_action__date=date_filtre)
    if user_filtre:
        qs = qs.filter(utilisateur__username__icontains=user_filtre)
    if modele_filtre:
        qs = qs.filter(modele__icontains=modele_filtre)

    # Stats par jour (7 derniers jours, dans l'ordre chronologique, jours à zéro inclus)
    today = timezone.now().date()
    jours = [today - timedelta(days=i) for i in range(6, -1, -1)]
    valeurs = [HistoriqueAction.objects.filter(date_action__date=j).count() for j in jours]
    stats_jour_json = {
        'labels': [j.strftime('%d/%m') for j in jours],
        'values': valeurs,
    }

    return render(request, 'apm/historique.html', {
        'actions': qs[:200],
        'stats_jour_json': json.dumps(stats_jour_json),
        'date_filtre': date_filtre,
        'user_filtre': user_filtre,
        'modele_filtre': modele_filtre,
        'total': qs.count(),
    })


# ===========================================================================
# DOCUMENTS — Upload de fichiers
# ===========================================================================

@login_required
def document_upload_fichier(request, pk):
    if not _dsi_allowed(request.user):
        messages.error(request, "Accès réservé à la DSI.")
        return redirect('documents_list')
    doc = get_object_or_404(Document, pk=pk)
    if request.method == 'POST':
        fichier = request.FILES.get('fichier')
        if not fichier:
            messages.error(request, "Aucun fichier sélectionné.")
            return redirect('documents_edit', pk=pk)

        # Validation extension
        ext = fichier.name.rsplit('.', 1)[-1].lower() if '.' in fichier.name else 'autre'
        EXTS_OK = {'pdf', 'docx', 'xlsx', 'png', 'jpg', 'jpeg'}
        if ext not in EXTS_OK:
            messages.error(request, f"Extension « .{ext} » non autorisée. Formats : PDF, DOCX, XLSX, PNG, JPG.")
            return redirect('documents_edit', pk=pk)

        type_map = {'pdf': 'pdf', 'docx': 'docx', 'xlsx': 'xlsx',
                    'png': 'png', 'jpg': 'jpg', 'jpeg': 'jpg'}
        taille_ko = max(1, fichier.size // 1024)

        df = DocumentFichier.objects.create(
            document=doc,
            fichier=fichier,
            nom_original=fichier.name,
            type_fichier=type_map.get(ext, 'autre'),
            taille_ko=taille_ko,
            uploade_par=request.user,
        )
        journaliser(request, f'Upload fichier : {fichier.name}', 'DocumentFichier', df,
                    details=f'Document : {doc.nom_fichier}')
        notifier(
            list(Utilisateur.objects.filter(role__type_role=Role.TYPE_MEMBRE_DSI)),
            'Nouveau fichier uploadé',
            f'Fichier « {fichier.name} » ajouté au document « {doc.nom_fichier} ».',
            type_notif='success', action_url=reverse('documents_edit', args=[pk])
        )
        messages.success(request, f'Fichier « {fichier.name} » uploadé ({taille_ko} Ko).')
    return redirect('documents_edit', pk=pk)


@login_required
def document_fichier_delete(request, pk):
    if not _dsi_allowed(request.user):
        messages.error(request, "Accès réservé à la DSI.")
        return redirect('documents_list')
    df = get_object_or_404(DocumentFichier, pk=pk)
    doc_pk = df.document.pk
    nom = df.nom_original
    if request.method == 'POST':
        df.fichier.delete(save=False)
        df.delete()
        journaliser(request, f'Suppression fichier : {nom}', 'DocumentFichier')
        messages.success(request, f'Fichier « {nom} » supprimé.')
    return redirect('documents_edit', pk=doc_pk)
